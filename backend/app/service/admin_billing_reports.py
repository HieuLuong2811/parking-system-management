from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.enums.parking import SubscriptionStatus
from app.models.users import Users
from app.models.subscriptions import UserSubscription
from app.models.plans import SubscriptionPlan
from app.models.terms import AcademicTerm
from app.models.roles import Roles
from app.models.user_roles import UserRoles

OVERDUE_REPORT_STATUSES = [
    SubscriptionStatus.PAYMENT_DUE,
    SubscriptionStatus.OVERDUE,
    SubscriptionStatus.CANCELED,
]


STATUS_VI = {
    "ACTIVE": "Đang hoạt động",
    "PAYMENT_DUE": "Đến hạn thanh toán",
    "OVERDUE": "Quá hạn thanh toán",
    "SUSPENDED": "Tạm ngưng",
    "CANCELED": "Đã hủy / còn công nợ",
    "INACTIVE": "Không hoạt động",
}


def format_status_vi(status: str | None) -> str:
    if not status:
        return "—"

    return STATUS_VI.get(str(status), str(status))


def get_user_phone(user: Users) -> str:
    return (
        getattr(user, "phone_number", None)
        or getattr(user, "phone", None)
        or "—"
    )

async def get_admin_emails(db: AsyncSession) -> list[str]:
    stmt = (
        select(Users.email)
        .join(UserRoles, UserRoles.user_code == Users.user_code)
        .join(Roles, Roles.id == UserRoles.role_id)
        .where(Roles.role_code == "admin")
        .where(Users.email.is_not(None))
    )

    result = await db.execute(stmt)

    emails = [
        str(email)
        for email in result.scalars().all()
        if email
    ]

    return sorted(set(emails))

async def get_unpaid_subscription_rows(db: AsyncSession) -> list[dict]:
    stmt = (
        select(
            UserSubscription,
            Users,
            SubscriptionPlan,
            AcademicTerm,
        )
        .join(Users, Users.user_code == UserSubscription.user_code)
        .join(SubscriptionPlan, SubscriptionPlan.id == UserSubscription.sub_plan_id)
        .join(AcademicTerm, AcademicTerm.id == UserSubscription.term_id)
        .where(UserSubscription.status.in_(OVERDUE_REPORT_STATUSES))
    )

    result = await db.execute(stmt)
    rows = []

    for subscription, user, plan, term in result.all():
        total_amount = float(subscription.total_amount or 0)
        paid_amount = float(subscription.paid_amount or 0)
        debt_amount = max(total_amount - paid_amount, 0)

        if debt_amount <= 0:
            continue

        rows.append(
            {
                "user_code": user.user_code,
                "full_name": user.full_name,
                "email": user.email,
                "phone_number": get_user_phone(user),
                "subscription_id": str(subscription.id),
                "plan_type": plan.plans_type,
                "term_name": term.term_name,
                "created_at": subscription.created_at,
                "status": format_status_vi(
                    subscription.status.value
                    if hasattr(subscription.status, "value")
                    else subscription.status
                ),
                "total_amount": total_amount,
                "paid_amount": paid_amount,
                "debt_amount": debt_amount,
            }
        )

    return rows


def build_unpaid_subscription_excel(rows: Iterable[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cong no goi gui xe"

    headers = [
        "Mã người dùng",
        "Họ tên người dùng",
        "Email",
        "Số điện thoại",
        "Mã vé xe đã đăng ký",
        "Loại vé xe",
        "Học kỳ",
        "Ngày đăng ký",
        "Trạng thái",
        "Tổng tiền",
        "Số tiền đã thanh toán",
        "Số tiền nợ",
    ]

    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        created_at = row["created_at"]
        if isinstance(created_at, datetime):
            created_at_value = created_at.strftime("%d/%m/%Y %H:%M")
        else:
            created_at_value = str(created_at or "—")

        sheet.append(
            [
                row["user_code"],
                row["full_name"],
                row["email"],
                row["phone_number"],
                row["subscription_id"],
                row["plan_type"],
                row["term_name"],
                created_at_value,
                row["status"],
                row["total_amount"],
                row["paid_amount"],
                row["debt_amount"],
            ]
        )

    money_columns = ["J", "K", "L"]

    for column in money_columns:
        for cell in sheet[column][1:]:
            cell.number_format = '#,##0" VND"'

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    widths = {
        "A": 16,
        "B": 24,
        "C": 30,
        "D": 18,
        "E": 40,
        "F": 18,
        "G": 20,
        "H": 20,
        "I": 24,
        "J": 18,
        "K": 22,
        "L": 18,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.read()